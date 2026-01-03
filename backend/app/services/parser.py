import logging
import pandas as pd
from fastapi import UploadFile, HTTPException
from io import BytesIO
from pathlib import Path
from app.core.config import get_settings
from app.core.sanitization import sanitize_filename, validate_column_name
from app.core.performance import track_performance
from openpyxl import load_workbook

logger = logging.getLogger(__name__)
settings = get_settings()


def unmerge_excel_cells(contents: bytes) -> pd.DataFrame:
    """
    Parse Excel file using openpyxl to handle merged cells.
    Merged cells are filled with the value from the top-left cell.
    
    Returns:
        DataFrame with merged cells properly filled
    """
    try:
        wb = load_workbook(BytesIO(contents), data_only=True)
        
        # Find largest sheet
        largest_sheet = None
        largest_rows = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row > largest_rows:
                largest_rows = ws.max_row
                largest_sheet = sheet_name
        
        ws = wb[largest_sheet] if largest_sheet else wb.active
        
        # Unmerge cells and fill with top-left value
        merged_ranges = list(ws.merged_cells.ranges)
        for merged_range in merged_ranges:
            # Get the value from the top-left cell
            top_left_cell = ws.cell(merged_range.min_row, merged_range.min_col)
            top_left_value = top_left_cell.value
            
            # Unmerge
            ws.unmerge_cells(str(merged_range))
            
            # Fill all cells in the range with the top-left value
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    ws.cell(row, col, top_left_value)
        
        if merged_ranges:
            logger.info(f"Unmerged {len(merged_ranges)} cell ranges in sheet '{largest_sheet}'")
        
        # Convert to DataFrame
        data = ws.values
        df = pd.DataFrame(data)
        
        # Use first row as header
        if len(df) > 0:
            df.columns = df.iloc[0]
            df = df[1:]
            df.reset_index(drop=True, inplace=True)
        
        return df
        
    except Exception as e:
        logger.warning(f"openpyxl parsing failed, falling back to pandas: {e}")
        return None  # Signal to use pandas fallback


# Allowed file extensions
ALLOWED_EXTENSIONS = {'.csv', '.xlsx', '.xls'}

# MIME type mapping for validation
MIME_TYPE_MAP = {
    'text/csv': '.csv',
    'application/vnd.ms-excel': '.xls',
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': '.xlsx',
}

def find_header_row(df: pd.DataFrame, max_scan_rows: int = 10) -> int:
    """
    Auto-detect the header row by finding the first row that looks like column headers.
    
    Heuristics:
    - Header rows have mostly unique, non-null string values
    - Header rows don't have mostly numeric values
    - Data rows below headers have consistent patterns
    
    Returns:
        Row index to use as header (0 = first row is header, as expected)
    """
    if len(df) < 2:
        return 0  # Not enough data to detect
    
    scan_limit = min(max_scan_rows, len(df))
    best_header_row = 0
    best_score = 0
    
    for row_idx in range(scan_limit):
        row = df.iloc[row_idx]
        
        # Score this row as a potential header
        score = 0
        non_null_count = row.notna().sum()
        
        if non_null_count == 0:
            continue  # Empty row, skip
            
        # Check if values are mostly strings (headers are usually text)
        string_count = sum(1 for v in row if isinstance(v, str) and len(str(v).strip()) > 0)
        string_ratio = string_count / non_null_count if non_null_count > 0 else 0
        
        # Check uniqueness (headers should be unique)
        unique_count = len(set(str(v).strip().lower() for v in row if pd.notna(v)))
        unique_ratio = unique_count / non_null_count if non_null_count > 0 else 0
        
        # Check if NOT mostly numeric (headers shouldn't be numbers)
        numeric_count = sum(1 for v in row if isinstance(v, (int, float)) and not pd.isna(v))
        numeric_ratio = numeric_count / non_null_count if non_null_count > 0 else 0
        
        # Scoring formula
        score = (string_ratio * 0.4) + (unique_ratio * 0.4) + ((1 - numeric_ratio) * 0.2)
        
        # Bonus for first row (default assumption)
        if row_idx == 0:
            score += 0.1
            
        if score > best_score:
            best_score = score
            best_header_row = row_idx
    
    return best_header_row


def validate_file_extension(filename: str) -> str:
    """
    Validate and sanitize file extension.
    Returns the extension if valid, raises HTTPException otherwise.
    """
    if not filename:
        raise HTTPException(status_code=400, detail="Filename is required")
    
    # Use pathlib for safe extension extraction
    file_path = Path(filename)
    file_ext = file_path.suffix.lower()
    
    if not file_ext:
        raise HTTPException(
            status_code=400,
            detail="File must have an extension. Supported formats: CSV, XLSX, XLS"
        )
    
    if file_ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file format: {file_ext}. Allowed formats: {', '.join(ALLOWED_EXTENSIONS)}"
        )
    
    return file_ext

def validate_mime_type(content_type: str, file_ext: str) -> None:
    """
    Validate MIME type matches file extension.
    Raises HTTPException if MIME type is suspicious.
    """
    if not content_type:
        # If no content type provided, rely on extension validation
        return
    
    # Check if MIME type matches expected extension
    expected_ext = MIME_TYPE_MAP.get(content_type.lower())
    if expected_ext and expected_ext != file_ext:
        # Log warning but don't fail for flexibility (some systems send wrong MIME types)
        logger.warning(f"MIME type {content_type} doesn't match extension {file_ext}")
    
    # Reject obviously malicious MIME types
    dangerous_types = [
        'application/x-executable',
        'application/x-sharedlib',
        'application/x-msdownload',
        'text/html',  # Could be XSS attempt
        'application/javascript',
    ]
    if content_type.lower() in dangerous_types:
        raise HTTPException(
            status_code=400,
            detail=f"File type '{content_type}' is not allowed. Only CSV and Excel files are supported."
        )

@track_performance("parse_file")
async def parse_file(file: UploadFile) -> pd.DataFrame:
    """
    Parse uploaded file into a pandas DataFrame.
    Validates file extension and MIME type before parsing.
    """
    try:
        # Validate file extension
        if not file.filename:
            raise HTTPException(status_code=400, detail="Filename is required")
        
        file_ext = validate_file_extension(file.filename)
        validate_mime_type(file.content_type, file_ext)  # Now raises on dangerous types
        
        # Read file contents
        contents = await file.read()
        
        if len(contents) == 0:
            raise HTTPException(status_code=400, detail="File is empty")
        
        # Parse based on extension
        if file_ext == '.csv':
            # Try different encodings
            try:
                df = pd.read_csv(BytesIO(contents), header=None)  # Read without header first
            except UnicodeDecodeError:
                try:
                    df = pd.read_csv(BytesIO(contents), encoding='latin1', header=None)
                except Exception as e:
                    logger.error(f"Error parsing CSV file: {e}")
                    raise HTTPException(
                        status_code=400,
                        detail="Unable to parse CSV file. Please ensure the file is properly formatted."
                    )
            
            # Auto-detect header row
            header_row = find_header_row(df)
            if header_row > 0:
                logger.info(f"Auto-detected header at row {header_row}, skipping {header_row} metadata rows")
            
            # Re-read with correct header
            df = pd.read_csv(BytesIO(contents), header=header_row) if header_row == 0 else pd.read_csv(BytesIO(contents), skiprows=range(header_row), header=0)
                
        elif file_ext in ['.xlsx', '.xls']:
            df = None
            
            # Try openpyxl first for better merged cell handling (xlsx only)
            if file_ext == '.xlsx':
                df = unmerge_excel_cells(contents)
            
            # Fallback to pandas if openpyxl failed or for .xls files
            if df is None:
                try:
                    # Read all sheets first to detect multi-sheet files
                    excel_file = pd.ExcelFile(BytesIO(contents))
                    sheet_names = excel_file.sheet_names
                    
                    if len(sheet_names) > 1:
                        # Multi-sheet file - auto-select largest sheet with data
                        largest_sheet = None
                        largest_rows = 0
                        
                        for sheet_name in sheet_names:
                            try:
                                test_df = pd.read_excel(excel_file, sheet_name=sheet_name)
                                if len(test_df) > largest_rows:
                                    largest_rows = len(test_df)
                                    largest_sheet = sheet_name
                            except:
                                continue
                        
                        if largest_sheet:
                            df = pd.read_excel(excel_file, sheet_name=largest_sheet)
                            logger.info(f"Multi-sheet Excel file detected. Selected '{largest_sheet}' ({largest_rows} rows) from {len(sheet_names)} sheets")
                        else:
                            # Fallback to first sheet
                            df = pd.read_excel(excel_file, sheet_name=sheet_names[0])
                    else:
                        # Single sheet - read normally
                        df = pd.read_excel(excel_file)
                except Exception as e:
                    logger.error(f"Error parsing Excel file: {e}")
                    raise HTTPException(
                        status_code=400,
                        detail="Unable to parse Excel file. Please ensure the file is not corrupted."
                    )
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format")
        
        if df.empty:
            raise HTTPException(
                status_code=400,
                detail="File appears to be empty or contains no data"
            )
            
        safe_filename = sanitize_filename(file.filename) if file.filename else 'unknown'
        logger.info(f"Successfully parsed file: {safe_filename}, shape: {df.shape}")
        return df
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Unexpected error parsing file {file.filename}: {e}", exc_info=True)
        raise HTTPException(
            status_code=400,
            detail="An error occurred while parsing the file. Please check the file format and try again."
        )

def validate_file_content(df: pd.DataFrame, filename: str) -> None:
    """
    Validate that the parsed file content is reasonable and safe.
    
    Args:
        df: Parsed DataFrame
        filename: Original filename for error messages
        
    Raises:
        HTTPException: If content validation fails
    """
    # Check for reasonable row/column limits to prevent memory exhaustion
    if len(df) > settings.max_file_rows:
        raise HTTPException(
            status_code=400,
            detail=f"File contains too many rows ({len(df):,}). Maximum allowed: {settings.max_file_rows:,} rows."
        )
    
    if len(df.columns) > settings.max_file_columns:
        raise HTTPException(
            status_code=400,
            detail=f"File contains too many columns ({len(df.columns)}). Maximum allowed: {settings.max_file_columns} columns."
        )
    
    # Validate column names
    for col in df.columns:
        if not validate_column_name(str(col)):
            raise HTTPException(
                status_code=400,
                detail=f"Invalid column name: '{col}'. Column names must be safe and not contain path traversal or special characters."
            )
    
    # Check for suspiciously large cell values (potential memory attack)
    for col in df.columns:
        if df[col].dtype == 'object':
            max_length = df[col].astype(str).str.len().max()
            if pd.notna(max_length) and max_length > settings.max_cell_size_bytes:
                raise HTTPException(
                    status_code=400,
                    detail=f"File contains extremely large text values in column '{col}'. Maximum allowed: {settings.max_cell_size_bytes} bytes per cell."
                )

def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    # Remove empty columns/rows
    df = df.dropna(how='all', axis=0)
    df = df.dropna(how='all', axis=1)
    
    # Sanitize column names - remove newlines, carriage returns, and excess whitespace
    # This is critical for Vega-Lite which breaks on newlines in field names
    def sanitize_column_name(col):
        if isinstance(col, str):
            # Replace newlines and carriage returns with space
            col = col.replace('\n', ' ').replace('\r', ' ')
            # Collapse multiple spaces
            col = ' '.join(col.split())
        return col
    
    df.columns = [sanitize_column_name(col) for col in df.columns]
    
    return df
